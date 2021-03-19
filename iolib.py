from enum import Enum
import struct
import json
from openpyxl import Workbook, load_workbook


class Endian(Enum):
    Little = 0
    Big = 1


class Excel:
    @staticmethod
    def pack(name: str, data: dict):
        wb = Workbook()
        ws = wb.active
        for r in range(len(data)):
            for c in range(len(data[r])):
                ws.cell(row=r+1, column=c+1, value=data[r][c])
        wb.save(name)

    @staticmethod
    def unpack(name: str) -> dict:
        wb = load_workbook(name)
        ws = wb.active
        arr = []
        for r in range(1, ws.max_row+1):
            arr2 = []
            for c in range(1, ws.max_column+1):
                arr2.append(ws.cell(row=r, column=c).value)
            arr.append(arr2)
        return arr


class Writer:
    def __init__(self, f, endian: Endian = Endian.Little):
        self.f = f
        if endian == Endian.Little:
            self.e = "<"
        else:
            self.e = ">"
    def close(self):
        self.f.close()
    def write(self, buf):
        self.f.write(buf)
    def write_8(self, data: int):
        buf = struct.pack(self.e + "b", data)
        self.write(buf)
        return 1
    def write_16(self, data: int):
        buf = struct.pack(self.e + "h", data)
        self.write(buf)
        return 2
    def write_32(self, data: int):
        buf = struct.pack(self.e + "i", data)
        self.write(buf)
        return 4
    def write_64(self, data: int):
        buf = struct.pack(self.e + "q", data)
        self.write(buf)
        return 8
    def write_u8(self, data: int):
        buf = struct.pack(self.e + "B", data)
        self.write(buf)
        return 1
    def write_u16(self, data: int):
        buf = struct.pack(self.e + "H", data)
        self.write(buf)
        return 2
    def write_u32(self, data: int):
        buf = struct.pack(self.e + "I", data)
        self.write(buf)
        return 4
    def write_u64(self, data: int):
        buf = struct.pack(self.e + "Q", data)
        self.write(buf)
        return 8
    def write_nt_utf8_str(self, data: str):
        buf = data.encode('utf-8')
        self.write(buf)
        self.write_8(0)
        return len(data) + 1
    def skip(self, count: int):
        for _ in range(count):
            self.write_8(0)
        return count
    def seek(self, offset: int):
        self.f.seek(offset)
    def tell(self):
        return self.f.tell()

class Reader:
    def __init__(self, f, endian: Endian = Endian.Little):
        self.f = f
        if endian == Endian.Little:
            self.e = "<"
        else:
            self.e = ">"
    def close(self):
        self.f.close()
    def read(self, count: int):
        return self.f.read(count)
    def read_8(self):
        buf = self.read(1)
        return struct.unpack(self.e + "b", buf)[0]
    def read_16(self):
        buf = self.read(2)
        return struct.unpack(self.e + "h", buf)[0]
    def read_32(self):
        buf = self.read(4)
        return struct.unpack(self.e + "i", buf)[0]
    def read_64(self):
        buf = self.read(8)
        return struct.unpack(self.e + "q", buf)[0]
    def read_u8(self):
        buf = self.read(1)
        return struct.unpack(self.e + "B", buf)[0]
    def read_u16(self):
        buf = self.f.read(2)
        return struct.unpack(self.e + "H", buf)[0]
    def read_u32(self):
        buf = self.read(4)
        return struct.unpack(self.e + "I", buf)[0]
    def read_u64(self):
        buf = self.read(8)
        return struct.unpack(self.e + "Q", buf)[0]
    def read_nt_utf8_str(self, sb=0):
        len = 0
        pos = self.tell()
        while self.read_8() != sb:
            len += 1
        self.seek(pos)
        line = self.read(len).decode('utf-8')
        self.skip(1)
        return line
    def seek(self, offset: int):
        self.f.seek(offset)
    def tell(self):
        return self.f.tell()
    def skip(self, count: int):
        self.seek(self.tell() + count)


class Stuff:
    @staticmethod
    def dump_json(name, data, indent: int = 0):
        with open(name, 'w', encoding='utf8') as json_file:
            json.dump(data, json_file, ensure_ascii=False, indent=indent)
            json_file.close()

    @staticmethod
    def load_json(name):
        with open(name, 'r', encoding='utf8') as json_file:
            return json.load(json_file)
